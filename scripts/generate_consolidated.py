#!/usr/bin/env python3
"""
종합 오더현황 Excel 생성 및 Google Drive 업로드 스크립트
Rachgia Dashboard v19 - 자동화 빌드 시스템

사용법:
    python scripts/generate_consolidated.py

입력: data/*.xlsx (4개 공장 파일)
출력: data/종합_오더현황_YYYY-MM-DD.xlsx → Google Drive 업로드

환경 변수 (업로드용, 없으면 로컬 생성만):
    GOOGLE_SERVICE_ACCOUNT_KEY: 서비스 계정 JSON 키
    GOOGLE_DRIVE_FOLDER_ID: Google Drive 폴더 ID
"""

import os
import sys
import json
import io
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages...")
    os.system("pip install pandas openpyxl")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# 프로젝트 루트를 sys.path에 추가 (parse_loadplan.py import용)
PROJECT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_DIR))
from parse_loadplan import parse_factory_file

# 설정
DATA_DIR = PROJECT_DIR / 'data'
FACTORY_FILES = {
    'A': 'Factory_A.xlsx',
    'B': 'Factory_B.xlsx',
    'C': 'Factory_C.xlsx',
    'D': 'Factory_D.xlsx',
}

# Excel 컬럼 정의 (31개)
EXCEL_COLUMNS = [
    'Factory', 'Unit', 'Season', 'Model', 'Article', 'Color',
    'Destination', 'PO Number', 'Quantity', 'CRD', 'CRD Month',
    'SDD', 'SDD Month', 'Code04', 'AQL', 'Inspection',
    'S_CUT', 'PRE_SEW', 'SEW_INPUT', 'SEW_BAL', 'S_FIT',
    'ASS_BAL', 'WH_IN', 'WH_OUT',
    'OSC Remaining', 'SEW Remaining', 'ASS Remaining',
    'WH_IN Remaining', 'WH_OUT Remaining',
    'Delay', 'Overall Status',
]


def flatten_record(record):
    """파싱된 레코드를 Excel 행으로 변환"""
    prod = record.get('production', {})
    remaining = record.get('remaining', {})

    # 지연 판정: SDD > CRD이고 Code04 승인이 없는 경우
    delay = False
    crd = record.get('crd', '')
    sdd = record.get('sddValue', '')
    code04 = record.get('code04')
    if crd and sdd and crd != '' and sdd != '':
        try:
            if sdd > crd and not code04:
                delay = True
        except (TypeError, ValueError):
            pass

    # Overall Status: WH_OUT 기준
    wh_out = prod.get('wh_out', {})
    overall_status = wh_out.get('status', 'pending')

    return [
        record.get('factory', ''),
        record.get('unit', ''),
        record.get('season', ''),
        record.get('model', ''),
        record.get('article', ''),
        record.get('color', ''),
        record.get('destination', ''),
        record.get('poNumber', ''),
        record.get('quantity', 0),
        crd,
        record.get('crdYearMonth', ''),
        sdd,
        record.get('sddYearMonth', ''),
        code04 or '',
        'Yes' if record.get('aql') else 'No',
        record.get('inspection') or '',
        # 공정별 완료수량
        prod.get('s_cut', {}).get('completed', 0),
        prod.get('pre_sew', {}).get('completed', 0),
        prod.get('sew_input', {}).get('completed', 0),
        prod.get('sew_bal', {}).get('completed', 0),
        prod.get('s_fit', {}).get('completed', 0),
        prod.get('ass_bal', {}).get('completed', 0),
        prod.get('wh_in', {}).get('completed', 0),
        prod.get('wh_out', {}).get('completed', 0),
        # 잔량
        remaining.get('osc', 0),
        remaining.get('sew', 0),
        remaining.get('ass', 0),
        remaining.get('whIn', 0),
        remaining.get('whOut', 0),
        # 지연/상태
        'Yes' if delay else 'No',
        overall_status,
    ]


def create_excel(all_records, output_path):
    """종합 오더현황 Excel 파일 생성"""
    wb = Workbook()

    # --- 종합 오더현황 시트 ---
    ws = wb.active
    ws.title = '종합 오더현황'

    # 스타일 정의
    header_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    delay_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    delay_font = Font(name='맑은 고딕', size=9, color='9C0006')
    data_font = Font(name='맑은 고딕', size=9)
    data_alignment = Alignment(vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')

    # 헤더 작성
    for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 작성
    for row_idx, record in enumerate(all_records, 2):
        row_data = flatten_record(record)
        is_delayed = row_data[29] == 'Yes'  # Delay 컬럼

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if is_delayed:
                cell.fill = delay_fill
                cell.font = delay_font
            else:
                cell.font = data_font

            # 숫자 컬럼 우측 정렬
            if col_idx in (9, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29):
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment

    # 컬럼 너비 자동 조정
    column_widths = {
        1: 8,   # Factory
        2: 8,   # Unit
        3: 10,  # Season
        4: 20,  # Model
        5: 15,  # Article
        6: 15,  # Color
        7: 12,  # Destination
        8: 15,  # PO Number
        9: 10,  # Quantity
        10: 12, # CRD
        11: 10, # CRD Month
        12: 12, # SDD
        13: 10, # SDD Month
        14: 10, # Code04
        15: 6,  # AQL
        16: 12, # Inspection
        17: 10, # S_CUT
        18: 10, # PRE_SEW
        19: 10, # SEW_INPUT
        20: 10, # SEW_BAL
        21: 10, # S_FIT
        22: 10, # ASS_BAL
        23: 10, # WH_IN
        24: 10, # WH_OUT
        25: 12, # OSC Remaining
        26: 12, # SEW Remaining
        27: 12, # ASS Remaining
        28: 13, # WH_IN Remaining
        29: 14, # WH_OUT Remaining
        30: 7,  # Delay
        31: 12, # Overall Status
    }
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 헤더 고정 (Freeze Panes)
    ws.freeze_panes = 'A2'

    # 자동 필터
    ws.auto_filter.ref = f'A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(all_records) + 1}'

    # --- Info 시트 ---
    ws_info = wb.create_sheet('Info')
    info_header_font = Font(name='맑은 고딕', bold=True, size=11, color='2F5496')
    info_font = Font(name='맑은 고딕', size=10)

    info_data = [
        ('생성 일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('총 오더 수', len(all_records)),
        ('', ''),
        ('공장별 오더 수', ''),
    ]

    # 공장별 집계
    factory_counts = {}
    for r in all_records:
        f = r.get('factory', 'Unknown')
        factory_counts[f] = factory_counts.get(f, 0) + 1

    for factory in sorted(factory_counts.keys()):
        info_data.append((f'  Factory {factory}', factory_counts[factory]))

    # 상태별 집계
    info_data.append(('', ''))
    info_data.append(('상태별 오더 수', ''))
    status_counts = {}
    for r in all_records:
        wh_out = r.get('production', {}).get('wh_out', {})
        status = wh_out.get('status', 'pending')
        status_counts[status] = status_counts.get(status, 0) + 1

    for status in ['completed', 'partial', 'pending']:
        if status in status_counts:
            info_data.append((f'  {status}', status_counts[status]))

    # 지연 오더 수
    delay_count = sum(1 for r in all_records if is_delayed_record(r))
    info_data.append(('', ''))
    info_data.append(('지연 오더 수', delay_count))

    for row_idx, (label, value) in enumerate(info_data, 1):
        cell_label = ws_info.cell(row=row_idx, column=1, value=label)
        cell_value = ws_info.cell(row=row_idx, column=2, value=value)

        if label and not label.startswith('  '):
            cell_label.font = info_header_font
        else:
            cell_label.font = info_font
        cell_value.font = info_font

    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 25

    # 저장
    wb.save(output_path)
    print(f"  ✅ Excel 파일 저장: {output_path}")
    print(f"     총 {len(all_records)}개 오더, {len(factory_counts)}개 공장")


def is_delayed_record(record):
    """지연 오더 여부 판정"""
    crd = record.get('crd', '')
    sdd = record.get('sddValue', '')
    code04 = record.get('code04')
    if crd and sdd and crd != '' and sdd != '':
        try:
            return sdd > crd and not code04
        except (TypeError, ValueError):
            pass
    return False


def get_drive_service():
    """Google Drive API 서비스 생성"""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        print("  ⚠️ Google API 라이브러리 없음 - 업로드 스킵")
        return None

    service_account_key = os.environ.get('GOOGLE_SERVICE_ACCOUNT_KEY')
    if not service_account_key:
        # 로컬 개발용
        key_file = PROJECT_DIR / 'credentials' / 'service-account.json'
        if key_file.exists():
            SCOPES = ['https://www.googleapis.com/auth/drive']
            creds = service_account.Credentials.from_service_account_file(
                str(key_file), scopes=SCOPES
            )
            return build('drive', 'v3', credentials=creds)
        print("  ⚠️ 서비스 계정 키 없음 - 업로드 스킵")
        return None

    SCOPES = ['https://www.googleapis.com/auth/drive']
    key_dict = json.loads(service_account_key)
    creds = service_account.Credentials.from_service_account_info(key_dict, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)


def find_existing_file(service, folder_id, filename):
    """Drive 폴더에서 동일 파일명 검색"""
    query = (
        f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    )
    results = service.files().list(
        q=query, pageSize=1, fields="files(id, name)"
    ).execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None


def upload_to_drive(service, folder_id, local_path, filename):
    """Google Drive에 파일 업로드 (기존 파일 있으면 덮어쓰기)"""
    from googleapiclient.http import MediaFileUpload

    media = MediaFileUpload(
        str(local_path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    existing_id = find_existing_file(service, folder_id, filename)

    if existing_id:
        # 기존 파일 업데이트 (파일 ID, 공유 설정 유지)
        file = service.files().update(
            fileId=existing_id,
            media_body=media,
        ).execute()
        print(f"  ✅ Drive 파일 업데이트: {filename} (ID: {existing_id})")
    else:
        # 새 파일 생성
        file_metadata = {
            'name': filename,
            'parents': [folder_id],
        }
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id',
        ).execute()
        print(f"  ✅ Drive 파일 생성: {filename} (ID: {file.get('id')})")

    return file


def main():
    """메인 실행"""
    print("=" * 60)
    print("📊 종합 오더현황 Excel 생성 시작")
    print(f"   시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 데이터 디렉토리 확인
    if not DATA_DIR.exists():
        print(f"❌ 데이터 디렉토리가 없습니다: {DATA_DIR}")
        sys.exit(1)

    # 공장별 파일 파싱
    all_records = []
    parsed_factories = []

    for factory, filename in FACTORY_FILES.items():
        filepath = DATA_DIR / filename
        if not filepath.exists():
            print(f"  ⚠️ 파일 없음: {filename}")
            continue

        records = parse_factory_file(factory, str(filepath))
        all_records.extend(records)
        parsed_factories.append(factory)
        print(f"  Factory {factory}: {len(records)}개 레코드")

    if not all_records:
        print("\n❌ 파싱된 데이터가 없습니다.")
        sys.exit(1)

    print(f"\n총 {len(all_records)}개 오더 ({', '.join(parsed_factories)})")

    # Excel 생성
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'종합_오더현황_{today}.xlsx'
    output_path = DATA_DIR / filename

    print(f"\n📄 Excel 파일 생성 중...")
    create_excel(all_records, output_path)

    # Google Drive 업로드
    folder_id = os.environ.get('GOOGLE_DRIVE_FOLDER_ID')
    if not folder_id:
        print(f"\n⚠️ GOOGLE_DRIVE_FOLDER_ID 미설정 - 업로드 스킵")
        print(f"   로컬 파일: {output_path}")
        return 0

    print(f"\n☁️ Google Drive 업로드 중...")
    service = get_drive_service()
    if not service:
        print(f"   로컬 파일만 생성됨: {output_path}")
        return 0

    try:
        upload_to_drive(service, folder_id, output_path, filename)
    except Exception as e:
        print(f"  ❌ 업로드 실패: {e}")
        print(f"   로컬 파일: {output_path}")
        return 1

    # 결과 요약
    print("\n" + "=" * 60)
    print("✅ 종합 오더현황 Excel 생성 및 업로드 완료!")
    print(f"   파일명: {filename}")
    print(f"   오더 수: {len(all_records)}개")
    print(f"   공장: {', '.join(parsed_factories)}")
    print("=" * 60)

    return 0


if __name__ == '__main__':
    sys.exit(main())
