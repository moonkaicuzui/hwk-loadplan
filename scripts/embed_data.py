#!/usr/bin/env python3
"""
Excel → JSON 변환 및 HTML 임베드 스크립트
Rachgia Dashboard v19 - 자동화 빌드 시스템

사용법:
    python scripts/embed_data.py

입력: data/*.xlsx (4개 공장 파일)
출력: rachgia_dashboard_v19.html (EMBEDDED_DATA 업데이트)
"""

import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

# 경로 설정
PROJECT_DIR = Path(__file__).parent.parent
DATA_DIR = PROJECT_DIR / 'data'
HTML_FILE = PROJECT_DIR / 'rachgia_dashboard_v19.html'

# 공장 파일 매핑
FACTORY_FILES = {
    'A': 'Factory_A.xlsx',
    'B': 'Factory_B.xlsx',
    'C': 'Factory_C.xlsx',
    'D': 'Factory_D.xlsx',
}

# 필수 컬럼 (Excel 헤더 → JSON 키)
COLUMN_MAPPING = {
    'order_no': ['Order No', 'ORDER NO', 'order_no', 'ORDER_NO', 'OrderNo'],
    'style': ['Style', 'STYLE', 'style', 'Style No', 'STYLE NO'],
    'destination': ['Destination', 'DESTINATION', 'destination', 'Dest', 'DEST'],
    'qty': ['Qty', 'QTY', 'qty', 'Quantity', 'QUANTITY', 'Order Qty'],
    'crd': ['CRD', 'crd', 'Customer Required Date'],
    'sdd': ['SDD', 'sdd', 'Scheduled Delivery Date', 'Ship Date'],
    's_cut': ['S_CUT', 's_cut', 'Cut', 'CUT', 'Cutting'],
    'pre_sew': ['PRE_SEW', 'pre_sew', 'Pre Sew', 'PRE SEW'],
    'sew_input': ['SEW_INPUT', 'sew_input', 'Sew Input', 'SEW INPUT'],
    'sew_bal': ['SEW_BAL', 'sew_bal', 'Sew Bal', 'SEW BAL', 'Sewing'],
    'osc': ['OSC', 'osc', 'Outsource'],
    'ass': ['ASS', 'ass', 'Assembly', 'ASSEMBLY'],
    'wh_in': ['WH_IN', 'wh_in', 'WH In', 'Warehouse In'],
    'wh_out': ['WH_OUT', 'wh_out', 'WH Out', 'Warehouse Out'],
    'aql': ['AQL', 'aql'],
    'code04': ['Code04', 'CODE04', 'code04', 'Code 04'],
}


def find_header_row(sheet):
    """헤더 행 찾기"""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_str = ' '.join(str(cell) for cell in row if cell)
        if any(keyword in row_str.upper() for keyword in ['ORDER', 'STYLE', 'QTY', 'CRD']):
            return row_idx
    return 1


def find_column_index(headers, target_keys):
    """컬럼 인덱스 찾기"""
    for idx, header in enumerate(headers):
        if header:
            header_str = str(header).strip()
            for key in target_keys:
                if header_str.lower() == key.lower():
                    return idx
    return None


def parse_date(value):
    """날짜 파싱"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        # 다양한 날짜 형식 처리
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y.%m.%d']:
            try:
                return datetime.strptime(value, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return str(value) if value else None


def parse_number(value):
    """숫자 파싱"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value) if value == int(value) else value
    if isinstance(value, str):
        try:
            cleaned = re.sub(r'[^\d.]', '', value)
            return int(float(cleaned)) if cleaned else 0
        except ValueError:
            return 0
    return 0


def parse_excel_file(file_path, factory_name):
    """Excel 파일 파싱"""
    print(f"\n📊 파싱 중: {file_path.name} (Factory {factory_name})")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    # 헤더 행 찾기
    header_row = find_header_row(sheet)
    headers = [cell for cell in sheet[header_row]]
    header_values = [cell.value for cell in headers]

    print(f"  헤더 행: {header_row}")
    print(f"  컬럼 수: {len(header_values)}")

    # 컬럼 인덱스 매핑
    col_indices = {}
    for json_key, possible_headers in COLUMN_MAPPING.items():
        idx = find_column_index(header_values, possible_headers)
        if idx is not None:
            col_indices[json_key] = idx

    print(f"  매핑된 컬럼: {list(col_indices.keys())}")

    # 데이터 파싱
    records = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        # 빈 행 스킵
        if not any(row):
            continue

        # Order No가 없으면 스킵
        order_no_idx = col_indices.get('order_no')
        if order_no_idx is None or not row[order_no_idx]:
            continue

        record = {
            'factory': factory_name,
        }

        for json_key, col_idx in col_indices.items():
            value = row[col_idx] if col_idx < len(row) else None

            # 타입별 파싱
            if json_key in ['crd', 'sdd']:
                record[json_key] = parse_date(value)
            elif json_key in ['qty', 's_cut', 'pre_sew', 'sew_input', 'sew_bal', 'osc', 'ass', 'wh_in', 'wh_out', 'aql']:
                record[json_key] = parse_number(value)
            else:
                record[json_key] = str(value).strip() if value else ''

        records.append(record)

    print(f"  파싱된 레코드: {len(records)}개")
    wb.close()

    return records


def update_html_with_data(html_path, all_data):
    """HTML 파일의 EMBEDDED_DATA 업데이트"""
    print(f"\n📝 HTML 업데이트 중: {html_path.name}")

    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # JSON 데이터 생성
    json_data = json.dumps(all_data, ensure_ascii=False, indent=2)

    # EMBEDDED_DATA 패턴 찾기 및 교체
    pattern = r'(const EMBEDDED_DATA = )\[[\s\S]*?\];'
    replacement = f'const EMBEDDED_DATA = {json_data};'

    if re.search(pattern, content):
        new_content = re.sub(pattern, replacement, content, count=1)
        print("  ✅ EMBEDDED_DATA 교체 완료")
    else:
        # 패턴이 없으면 삽입 위치 찾기
        insert_pattern = r'(<script>[\s\S]*?// =+ Data =+)'
        if re.search(insert_pattern, content):
            new_content = re.sub(
                insert_pattern,
                f'\\1\n        {replacement}',
                content,
                count=1
            )
            print("  ✅ EMBEDDED_DATA 삽입 완료")
        else:
            print("  ❌ EMBEDDED_DATA 삽입 위치를 찾을 수 없습니다.")
            return False

    # 버전 업데이트
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    version_pattern = r"(CACHE_VERSION = 'rachgia-v)[\d.]+(';)"

    # 버전 번호 추출 및 증가
    version_match = re.search(r"CACHE_VERSION = 'rachgia-v([\d.]+)'", content)
    if version_match:
        old_version = version_match.group(1)
        parts = old_version.split('.')
        parts[-1] = str(int(parts[-1]) + 1)
        new_version = '.'.join(parts)
        new_content = re.sub(version_pattern, f"\\g<1>{new_version}\\2", new_content)
        print(f"  버전 업데이트: v{old_version} → v{new_version}")

    # 파일 저장
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    print(f"  ✅ 저장 완료: {html_path}")
    return True


def main():
    """메인 실행 함수"""
    print("=" * 60)
    print("🚀 Excel → HTML 임베드 시작")
    print(f"   시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 데이터 디렉토리 확인
    if not DATA_DIR.exists():
        print(f"❌ 데이터 디렉토리가 없습니다: {DATA_DIR}")
        print("   먼저 download_from_drive.py를 실행하세요.")
        sys.exit(1)

    # Excel 파일 파싱
    all_data = []
    parsed_factories = []

    for factory_name, file_name in FACTORY_FILES.items():
        file_path = DATA_DIR / file_name

        if not file_path.exists():
            print(f"\n⚠️ 파일 없음: {file_name}")
            continue

        try:
            records = parse_excel_file(file_path, factory_name)
            all_data.extend(records)
            parsed_factories.append(factory_name)
        except Exception as e:
            print(f"  ❌ 파싱 오류: {e}")

    if not all_data:
        print("\n❌ 파싱된 데이터가 없습니다.")
        sys.exit(1)

    # HTML 업데이트
    if not HTML_FILE.exists():
        print(f"\n❌ HTML 파일이 없습니다: {HTML_FILE}")
        sys.exit(1)

    success = update_html_with_data(HTML_FILE, all_data)

    # 결과 요약
    print("\n" + "=" * 60)
    print("✅ 임베드 완료!")
    print(f"   공장: {', '.join(parsed_factories)}")
    print(f"   총 레코드: {len(all_data)}개")
    print("=" * 60)

    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
